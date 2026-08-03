"""
session_template.py

Генератор xlsx-пресета сессии Linken Sphere из шаблона.
Архитектурно аналогичен sample_urls_for_run() из warmup_api.py:
stateless функция, читает шаблон → рандомит fingerprint-поля →
сохраняет per-machine файл.

Что фиксировано:
    System=windows, Browser=chrome, Connection=direct, WebRTC=fake
    DNS=1.1.1.1;1.0.0.1 (Cloudflare)
    Canvas/WebGL/ClientRects/Audio/WebGPU/MediaDevices = все fake
    Description, Import cookies, Import passwords = пусто

Что рандомится per-machine:
    System version: 10 / 11
    Video adapter:  21 позиция строго из каталога LS для Windows, с уклоном
                    в бюджетные ноутбуки 2018-2024 (Celeron/Pentium/Athlon,
                    массовый офис на UHD 620 / Vega, бюджетные игровые
                    GTX 1050/1650 и RTX 3050 Laptop) плюс офисные и игровые
                    десктопы.
    CPU / RAM / Screen: выбираются ИЗ ПРОФИЛЯ выбранной видеокарты
                       (см. ADAPTER_PROFILES). Так исключены абсурдные
                       сборки типа RTX 4070 + 4 ядра + 8 GB + 1080p.

Колонки строки 3 в шаблоне:
    A=session_name  B=description     C=connection      D=ip
    E=port          F=login           G=password        H=webrtc
    I=change_ip_url J=custom_dns      K=system          L=browser
    M=cpu           N=ram             O=screen          P=system_version
    Q=video_adapter R=canvas          S=webgl           T=client_rects
    U=audio         V=webgpu          W=media_devices   X=import_cookies
    Y=import_passwords
"""
from __future__ import annotations

import random
from pathlib import Path

import openpyxl


# --- фиксированные значения (одинаково для всех машин) ---
FIXED = {
    "B": "",                 # description пусто
    "C": "direct",           # connection protocol
    "H": "fake",             # webrtc
    "J": "1.1.1.1;1.0.0.1",  # custom DNS — Cloudflare
    "K": "windows",          # system
    "L": "chrome",           # browser
    "R": "fake",             # canvas
    "S": "fake",             # webgl   (было "noise" в шаблоне)
    "T": "fake",             # client rects (было "noise")
    "U": "direct",           # audio — OfflineAudioContext генерит в RAM, не лезет в
                             #         аппаратную звуковую подсистему → на Server 2022
                             #         без звуковой карты direct БЕЗОПАСЕН и неотличим
                             #         от обычного Chromium-on-Windows. LS-индикатор
                             #         показывает "close to real" против "too unique"
                             #         у noise/fake. Кроме того, noise детектируется
                             #         repeat-call тестом (Safari 17 публично описал
                             #         методику) — анти-фрод видит что hash плавает
                             #         между обращениями и игнорит fingerprint.
    "V": "fake",             # webgpu
    "W": "fake",             # media devices
    "X": "",                 # import cookies пусто
    "Y": "",                 # import passwords пусто
}

SYS_VERSIONS = ["10", "11"]

# Профили sane-комбинаций: для каждой видеокарты — допустимые CPU/RAM/Screen.
# Названия строго из официального каталога LS (лист "Инструкция RU", ячейка Q7
# в _template.xlsx; разрешения — O7, CPU — M7, RAM — N7). Любое расхождение
# хотя бы в символе → Mass Import отвергнет строку.
#
# ЧТО ОЗНАЧАЕТ ПОЛЕ CPU: браузер отдаёт navigator.hardwareConcurrency, то есть
# ЛОГИЧЕСКИЕ процессоры (потоки), а не физические ядра. Отсюда значения:
#   2  — Celeron 2c/2t             8  — 4c/8t (i5-8250U, Ryzen 5 3500U) ← самый частый ноут
#   4  — i3/Pentium 2c/4t, 4c/4t   12 — 6c/12t (i5-11400H, Ryzen 5 5500U)
#   6  — 6c/6t: типично для ДЕСКТОПОВ (i5-9400F), в ноутбуках почти не бывает
#
# ЭКРАНЫ. Доля 2560x1440 намеренно держится около реальной статистики (~12%):
# 1440p оставлен только там, где он оправдан — игровые/производительные
# десктопные карты. Все ноутбучные позиции и офисные iGPU — FHD, потому что
# в жизни офисный ПК и массовый ноут почти всегда 1920x1080.
#
# Состав пула намеренно смещён в сторону бюджетных ноутбуков 2018-2024:
# именно такие машины преобладают у реальных людей.
ADAPTER_PROFILES: dict[str, tuple[list[int], list[int], list[str]]] = {
    # ================= НОУТБУКИ: ультрабюджет 2018-2022 =================
    # Celeron / Pentium Silver / Athlon. 4 GB RAM тут ещё массово встречается.
    "Intel, UHD Graphics 600":                       ([2],     [4, 8],  ["1920x1080"]),
    "Intel, UHD Graphics 605":                       ([4],     [4, 8],  ["1920x1080"]),
    "AMD, Radeon(TM) Vega 3 Graphics":               ([2, 4],  [4, 8],  ["1920x1080"]),

    # ================= НОУТБУКИ: массовый офис 2018-2021 =================
    # UHD 620 (i3/i5-8xxxU) — самый распространённый офисный ноут своей эпохи.
    "Intel, UHD Graphics 620":                       ([4, 8],  [8, 16], ["1920x1080"]),
    "AMD, Radeon(TM) Vega 8 Mobile Graphics":        ([4, 8],  [8, 16], ["1920x1080"]),
    "AMD, Radeon(TM) Vega 10 Mobile Graphics":       ([8],     [8, 16], ["1920x1080"]),

    # ================= НОУТБУКИ: бюджетный игровой 2018-2024 =============
    # 1050/1050 Ti — массовый игровой 2018-2019; 1650/1650 Ti — 2019-2022;
    # RTX 3050 Laptop — 2021-2024. Все FHD: 1440p-панели в этом классе редкость.
    "Nvidia, GeForce GTX 1050":                      ([4, 8],  [8, 16], ["1920x1080"]),
    "Nvidia, GeForce GTX 1050 Ti":                   ([8],     [8, 16], ["1920x1080"]),
    "Nvidia, GeForce GTX 1650":                      ([8, 12], [8, 16], ["1920x1080"]),
    "Nvidia, GeForce GTX 1650 Ti":                   ([8, 12], [8, 16], ["1920x1080"]),
    "Nvidia, GeForce RTX 3050 Laptop GPU":           ([8, 12], [8, 16], ["1920x1080"]),

    # ================= НОУТБУКИ: средний игровой 2019-2023 ===============
    # Топовые (RTX 3080/4080 Laptop) намеренно НЕ включены — редкие машины,
    # их присутствие в пуле выглядело бы неестественно частым.
    "Nvidia, GeForce GTX 1660 Ti with Max-Q Design": ([12],    [16],    ["1920x1080"]),
    "Nvidia, GeForce RTX 3050 Ti Laptop GPU":        ([12],    [16],    ["1920x1080"]),

    # ================= ДЕСКТОПЫ: офисные Intel iGPU =====================
    # Офисный ПК на встроенной графике — почти всегда FHD-монитор.
    "Intel, UHD Graphics 770":                       ([4, 6, 8], [8, 16], ["1920x1080"]),
    "Intel, UHD Graphics 630":                       ([4, 6, 8], [8, 16], ["1920x1080"]),
    "Intel, Iris(R) Xe Graphics":                    ([4, 6, 8], [8, 16], ["1920x1080"]),
    "Intel, UHD Graphics":                           ([4, 6, 8], [8, 16], ["1920x1080"]),

    # ================= ДЕСКТОПЫ: игровые =================================
    # Здесь 1440p уместен — под такие карты берут 2K-монитор.
    # GTX 1660 (2019-2021): с i3/Ryzen 3 бывает крайне редко → от 6 потоков.
    "Nvidia, GeForce GTX 1660":                      ([6, 8],  [8, 16], ["1920x1080", "2560x1440"]),
    # RTX 3060 / RX 6600 (2021+): 8 GB с современной картой — редкость.
    "Nvidia, GeForce RTX 3060":                      ([6, 8],  [16],    ["1920x1080", "2560x1440"]),
    "AMD, Radeon RX 6600":                           ([6, 8],  [16],    ["1920x1080", "2560x1440"]),
    # RTX 4070 (2023+): ставят к i7/Ryzen 7+, 1080p с такой картой — мисматч.
    "Nvidia, GeForce RTX 4070":                      ([8],     [16],    ["2560x1440"]),
}


def build_sessions_xlsx(
    template: Path,
    target: Path,
    session_names: list[str],
    *,
    rng: random.Random | None = None,
) -> Path:
    """Сгенерировать per-machine xlsx из шаблона — N сессий одним файлом.

    Mass Import в LS читает строки начиная с 3-й: одна строка = одна
    сессия. Пишем len(session_names) строк (3, 4, 5, …), у КАЖДОЙ —
    свой независимый рандомный fingerprint (видеокарта → CPU/RAM/Screen
    из её профиля + версия Windows). Два профиля с одной VPS не должны
    выглядеть как один и тот же "компьютер".

    Args:
        template: путь к session_imports/_template.xlsx
        target: куда сохранить готовый файл
        session_names: имена сессий (по одному на строку, порядок сохраняется)
        rng: опциональный random.Random для тестов / детерминизма.
             По умолчанию использует системный random (true random).

    Returns:
        Path == target (для chain-style вызова).
    """
    if rng is None:
        rng = random.Random()

    if not session_names:
        raise ValueError("session_names пуст — нечего генерить")

    if not template.exists():
        raise FileNotFoundError(f"шаблон не найден: {template}")

    wb = openpyxl.load_workbook(template)
    ws = wb[wb.sheetnames[0]]

    for i, session_name in enumerate(session_names):
        row = 3 + i
        ws[f"A{row}"] = session_name

        for cell, val in FIXED.items():
            ws[f"{cell}{row}"] = val

        # Сначала выбираем видеокарту — она определяет реалистичный набор
        # CPU / RAM / Screen из своего профиля. Это исключает абсурдные
        # сочетания (RTX 4070 с 4 ядрами и т.п.).
        adapter = rng.choice(list(ADAPTER_PROFILES.keys()))
        cpu_options, ram_options, screen_options = ADAPTER_PROFILES[adapter]

        ws[f"M{row}"] = rng.choice(cpu_options)
        ws[f"N{row}"] = rng.choice(ram_options)
        ws[f"O{row}"] = rng.choice(screen_options)
        ws[f"P{row}"] = rng.choice(SYS_VERSIONS)
        ws[f"Q{row}"] = adapter

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target


def build_session_xlsx(
    template: Path,
    target: Path,
    session_name: str,
    *,
    rng: random.Random | None = None,
) -> Path:
    """Одиночная сессия — тонкая обёртка над build_sessions_xlsx.
    Оставлена для обратной совместимости вызовов."""
    return build_sessions_xlsx(template, target, [session_name], rng=rng)


if __name__ == "__main__":
    # smoke-тест: показать 10 рандомных распределений
    import sys
    root = Path(__file__).resolve().parent
    tpl = root / "session_imports" / "_template.xlsx"
    if len(sys.argv) > 1:
        tpl = Path(sys.argv[1])

    print(f"шаблон: {tpl}")
    print(f"генерим 10 примеров:\n")
    for i in range(10):
        out = Path(f"/tmp/test_session_{i}.xlsx")
        build_session_xlsx(tpl, out, f"test-{i:02d}")
        wb = openpyxl.load_workbook(out)
        ws = wb[wb.sheetnames[0]]
        print(
            f"  [{i:02d}] sysver=Win{ws['P3'].value}  CPU={ws['M3'].value}c  "
            f"RAM={ws['N3'].value}GB  screen={ws['O3'].value}  "
            f"video={ws['Q3'].value}"
        )
