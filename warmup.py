"""
Linken Sphere 2 — автоматизация Warm up.

Сценарий:
  0. Запустить Linken Sphere 2 если ещё не открыт, дождаться главного экрана.
  1. Кликнуть кнопку "три точки" → пункт "Warm up".
  2. Изменить viewing depth и time per url на значения из config.ini.
  3. Снять галочку "use most popular".
  4. Browse file → выбрать случайный .txt из указанной папки.
  5. Удалить первые N строк из URL-textarea (опционально).
  6. Нажать START.

Поиск элементов — по template matching (картинки в templates/).
Диалог открытия файла — нативный Windows, обрабатывается через клавиатуру.
"""

from __future__ import annotations

import configparser
import ctypes
import json
import logging
import os
import random
import subprocess
import sys
import time
from pathlib import Path

# DPI awareness ОБЯЗАТЕЛЬНО до импорта pyautogui/Pillow.
# Иначе на Windows со scaling != 100% ImageGrab отдаёт физические пиксели,
# а pyautogui.click трактует их как логические — клики уходят не туда.
if sys.platform == "win32":
    for _setter in (
        lambda: ctypes.windll.user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4)),
        lambda: ctypes.windll.shcore.SetProcessDpiAwareness(2),
        lambda: ctypes.windll.user32.SetProcessDPIAware(),
    ):
        try:
            _setter()
            break
        except (AttributeError, OSError):
            continue

import cv2
import numpy as np
import pyautogui
from PIL import ImageDraw, ImageGrab

# Помечает последние N точек клика на скриншоте, чтобы было видно куда летели курсоры.
_CLICK_TRAIL: list[tuple[int, int, str]] = []
_TRAIL_LIMIT = 6

ROOT = Path(__file__).resolve().parent
TEMPLATES_DIR = ROOT / "templates"
SCREENSHOTS_DIR = ROOT / "screenshots"
LOG_FILE = ROOT / "warmup.log"

pyautogui.FAILSAFE = True  # двинуть мышь в угол — аварийный стоп
pyautogui.PAUSE = 0.15

# Push-уведомления через ntfy.sh: на каждой машине без настройки.
# Топик играет роль «адреса» — кто знает строку, может слать push в эту тему.
# Подписаться: приложение ntfy → Subscribe to topic → ввести значение ниже.
NTFY_TOPIC = "warmup-r2d2-7m9k4n2p8q5xFx168xx1QQE"
# Сколько первых УСПЕШНЫХ запусков на новой машине ещё подтверждаем push'ем —
# чтобы убедиться, что setup отработал. Дальше — тишина (только при падениях).
SUCCESS_NOTIFY_COUNT = 2
SUCCESS_STATE_FILE = ROOT / ".warmup_state"
# Сгенерированное имя сессии этой машины (формат CL-XXXXXXXX, 8 цифр).
# Создаётся один раз — на первой инсталляции — и больше не меняется.
SESSION_NAME_FILE = ROOT / ".session_name"
# Флаг, что в LS активирован API-порт (Settings → Network → Api port).
# Если флаг есть — UI-активацию пропускаем, дальше всё через HTTP.
API_ACTIVATED_FLAG = ROOT / ".api_activated"
# Флаг, что сессия уже импортирована в LS на этой машине (хранит имя).
# Первый запуск импортит xlsx из session_imports/, дальше только warmup.
SESSION_IMPORTED_FLAG = ROOT / ".session_imported"
SESSION_IMPORTS_DIR = ROOT / "session_imports"


def setup_logging() -> logging.Logger:
    logger = logging.getLogger("warmup")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)
    return logger


log = setup_logging()


def load_config() -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    cfg_path = ROOT / "config.ini"
    if not cfg_path.exists():
        log.error("config.ini не найден рядом со скриптом")
        sys.exit(2)
    cfg.read(cfg_path, encoding="utf-8")
    return cfg


def _record_click(x: int, y: int, label: str) -> None:
    _CLICK_TRAIL.append((x, y, label))
    if len(_CLICK_TRAIL) > _TRAIL_LIMIT:
        _CLICK_TRAIL.pop(0)


def screenshot(step: str, enabled: bool) -> None:
    if not enabled:
        return
    SCREENSHOTS_DIR.mkdir(exist_ok=True)
    ts = time.strftime("%Y%m%d-%H%M%S")
    path = SCREENSHOTS_DIR / f"{ts}_{step}.png"
    try:
        img = ImageGrab.grab()
    except OSError as e:
        # screen grab падает в отключённой RDP-сессии — не валим из-за этого
        # ВЕСЬ скрипт, тем более в except-блоке при логировании ошибки.
        log.warning("screenshot не сделан (%s) — пропускаю", e)
        return
    if _CLICK_TRAIL:
        draw = ImageDraw.Draw(img)
        r = 18
        for x, y, label in _CLICK_TRAIL:
            draw.ellipse([x - r, y - r, x + r, y + r], outline="red", width=3)
            draw.line([(x - r, y), (x + r, y)], fill="red", width=2)
            draw.line([(x, y - r), (x, y + r)], fill="red", width=2)
            draw.text((x + r + 3, y - r), label, fill="red")
    img.save(path)
    log.info("screenshot → %s", path.name)


def grab_screen_bgr() -> np.ndarray:
    img = np.array(ImageGrab.grab())
    return cv2.cvtColor(img, cv2.COLOR_RGB2BGR)


def find_template(name: str, confidence: float) -> tuple[int, int] | None:
    """Возвращает центр найденного шаблона на экране или None."""
    tpl_path = TEMPLATES_DIR / f"{name}.png"
    if not tpl_path.exists():
        log.error("шаблон не найден: %s", tpl_path)
        return None

    tpl = cv2.imread(str(tpl_path), cv2.IMREAD_COLOR)
    if tpl is None:
        log.error("не удалось прочитать %s", tpl_path)
        return None

    screen = grab_screen_bgr()
    res = cv2.matchTemplate(screen, tpl, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, max_loc = cv2.minMaxLoc(res)
    log.info("match %s: confidence=%.3f", name, max_val)
    if max_val < confidence:
        return None
    h, w = tpl.shape[:2]
    return (max_loc[0] + w // 2, max_loc[1] + h // 2)


def wait_for(name: str, confidence: float, timeout: float) -> tuple[int, int]:
    deadline = time.time() + timeout
    while time.time() < deadline:
        pt = find_template(name, confidence)
        if pt is not None:
            return pt
        time.sleep(0.5)
    raise TimeoutError(f"элемент '{name}' не появился за {timeout}s")


def click(name: str, cfg: configparser.ConfigParser, *, double: bool = False) -> None:
    conf = cfg.getfloat("matching", "confidence")
    timeout = cfg.getfloat("matching", "wait_seconds")
    x, y = wait_for(name, conf, timeout)
    log.info("click %s@(%d,%d)%s", name, x, y, " ×2" if double else "")
    _record_click(x, y, name)
    if double:
        pyautogui.doubleClick(x, y)
    else:
        pyautogui.click(x, y)
    time.sleep(cfg.getfloat("matching", "step_delay"))


def click_at_offset(name: str, xf: float, yf: float, cfg: configparser.ConfigParser) -> None:
    """Кликает в точку (xf, yf) внутри найденного шаблона, доли 0..1 от ширины/высоты."""
    left, top, w, h = _find_template_box(name, cfg)
    x, y = int(left + w * xf), int(top + h * yf)
    log.info("click %s @offset(%.2f,%.2f) → (%d,%d)", name, xf, yf, x, y)
    _record_click(x, y, name)
    pyautogui.click(x, y)
    time.sleep(cfg.getfloat("matching", "step_delay"))


def _find_template_box(name: str, cfg: configparser.ConfigParser) -> tuple[int, int, int, int]:
    """Возвращает (left, top, width, height) найденного шаблона."""
    conf = cfg.getfloat("matching", "confidence")
    timeout = cfg.getfloat("matching", "wait_seconds")
    tpl_path = TEMPLATES_DIR / f"{name}.png"
    tpl = cv2.imread(str(tpl_path), cv2.IMREAD_COLOR)
    if tpl is None:
        raise FileNotFoundError(f"шаблон {tpl_path} не читается")
    h, w = tpl.shape[:2]

    deadline = time.time() + timeout
    while time.time() < deadline:
        res = cv2.matchTemplate(grab_screen_bgr(), tpl, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, max_loc = cv2.minMaxLoc(res)
        if max_val >= conf:
            log.info("match %s: confidence=%.3f", name, max_val)
            return (max_loc[0], max_loc[1], w, h)
        time.sleep(0.5)
    raise TimeoutError(f"элемент '{name}' не появился за {timeout}s")


def set_stepper(template_name: str, target: int, minimum: int, cfg: configparser.ConfigParser) -> None:
    """
    Stepper-поле: сначала кликаем '-' много раз (сброс к минимуму),
    затем '+' нужное число раз. Координаты +/- относительно box'а шаблона.
    """
    left, top, w, h = _find_template_box(template_name, cfg)
    plus_xf = cfg.getfloat("stepper_offsets", "plus_x")
    minus_xf = cfg.getfloat("stepper_offsets", "minus_x")
    yf = cfg.getfloat("stepper_offsets", "y_center")
    plus_pt = (int(left + w * plus_xf), int(top + h * yf))
    minus_pt = (int(left + w * minus_xf), int(top + h * yf))
    reset_n = cfg.getint("warmup", "reset_clicks")
    delta = max(0, target - minimum)

    log.info("stepper %s: reset to min via %d × '-' @%s", template_name, reset_n, minus_pt)
    _record_click(*minus_pt, f"{template_name}-")
    pyautogui.click(minus_pt[0], minus_pt[1], clicks=reset_n, interval=0.1)
    time.sleep(0.5)  # дать UI устаканиться после серии "-"

    log.info("stepper %s: increment %d × '+' @%s → target=%d", template_name, delta, plus_pt, target)
    _record_click(*plus_pt, f"{template_name}+")
    if delta > 0:
        pyautogui.click(plus_pt[0], plus_pt[1], clicks=delta, interval=0.15)
    time.sleep(cfg.getfloat("matching", "step_delay"))


def remove_first_lines_from_list(n: int, cfg: configparser.ConfigParser) -> None:
    """
    Удаляет первые n строк из URL-textarea в окне Warm up.
    Якоримся на browse_file_button, кликаем со сдвигом внутрь списка,
    Ctrl+Home → Shift+Down × n → Delete.
    """
    if n <= 0:
        return
    left, top, w, h = _find_template_box("browse_file_button", cfg)
    bx = left + w // 2
    by = top + h // 2
    dx = cfg.getint("url_list_offset", "dx_from_browse")
    dy = cfg.getint("url_list_offset", "dy_from_browse")
    click_x, click_y = bx + dx, by + dy

    log.info("removing first %d lines: click URL list @(%d,%d)", n, click_x, click_y)
    _record_click(click_x, click_y, "url_list")
    pyautogui.click(click_x, click_y)
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "home")
    time.sleep(0.15)
    pyautogui.keyDown("shift")
    for _ in range(n):
        pyautogui.press("down")
        time.sleep(0.04)
    pyautogui.keyUp("shift")
    time.sleep(0.15)
    pyautogui.press("delete")
    time.sleep(cfg.getfloat("matching", "step_delay"))


def load_credentials() -> configparser.ConfigParser:
    """Читает credentials.ini рядом со скриптом."""
    creds = configparser.ConfigParser()
    path = ROOT / "credentials.ini"
    if not path.exists():
        raise FileNotFoundError(
            "credentials.ini не найден. "
            "Скопируй credentials.ini.example в credentials.ini и заполни данные."
        )
    creds.read(path, encoding="utf-8")
    return creds


def _read_success_count() -> int:
    try:
        return int(SUCCESS_STATE_FILE.read_text(encoding="utf-8").strip())
    except (FileNotFoundError, ValueError):
        return 0


def _write_success_count(n: int) -> None:
    try:
        SUCCESS_STATE_FILE.write_text(str(n), encoding="utf-8")
    except OSError as e:
        log.warning("не удалось записать %s: %s", SUCCESS_STATE_FILE, e)


# Кэш публичного IP — один HTTP-запрос на Python-процесс, потом в памяти.
_machine_ip_cache: str | None = None


def _machine_id() -> str:
    """Публичный IP машины — стабильный уникальный идентификатор VPS.
    Hostname часто бесполезен (по дефолту 'WIN-XXX' или одинаковый
    'AKOPTO' у админ-аккаунта). IP уникален у каждой VPS.

    Кэшируется в in-process переменной на всё время жизни Python-процесса.
    На каждом новом Python-запуске пере-фетчится (provider мог переназначить
    IP после ребута). ipify основной, checkip.amazonaws.com fallback —
    оба отдают plain text IP в теле, без заголовков. Если сеть лежит —
    возвращаем 'no-ip', юзер сразу видит косяк."""
    global _machine_ip_cache
    if _machine_ip_cache:
        return _machine_ip_cache
    import urllib.request
    for url in ("https://api.ipify.org", "https://checkip.amazonaws.com"):
        try:
            with urllib.request.urlopen(url, timeout=5) as resp:
                ip = resp.read().decode("ascii", errors="ignore").strip()
                if ip:
                    _machine_ip_cache = ip
                    return ip
        except Exception:
            continue
    return "no-ip"


def _ntfy_header() -> str:
    """Единый префикс для всех ntfy-сообщений: session + machine.
    session ВСЕГДА первым — на телефоне его удобно ловить глазом,
    т.к. имена машин (hostname) могут совпадать между VPS."""
    try:
        sess = load_session_name()
    except Exception:
        sess = "<unknown>"
    return f"session: {sess}\nmachine: {_machine_id()}\n"


# Эмодзи в Title по типу события (по первому тегу). Telegram-бридж НЕ
# подставляет эмодзи из ntfy-тегов, поэтому кладём их прямо в заголовок.
_TAG_EMOJI = {"white_check_mark": "✅", "tada": "🎉", "warning": "⚠️"}
_PRIORITY_NUM = {"min": 1, "low": 2, "default": 3, "high": 4, "max": 5, "urgent": 5}


def notify_ntfy(message: str, title: str = "warmup failed",
                priority: str = "high", tags: str = "warning") -> None:
    """Шлёт push через ntfy.sh JSON-публикацией, без настройки на машине.
    Все ошибки (сети нет, сервис лежит, и т.д.) глотаются — нотификация
    никогда не должна мешать основному логированию.

    Дефолт priority/tags = high/warning рассчитан на сообщения о падении.
    Для успешных уведомлений вызывающий передаёт low/white_check_mark.

    JSON, а не HTTP-заголовки: в заголовки нельзя положить эмодзи (только
    ASCII), поэтому Title в Telegram приходил без иконки. В JSON-теле
    json.dumps экранирует Unicode в \\uXXXX — тело уходит чистым ASCII."""
    try:
        import urllib.request
        tag_list = [t.strip() for t in tags.split(",") if t.strip()]
        emoji = _TAG_EMOJI.get(tag_list[0], "") if tag_list else ""
        disp_title = f"{emoji} {title}".strip()
        payload = {
            "topic": NTFY_TOPIC,
            "title": disp_title,
            "message": message[:4000],
            "priority": _PRIORITY_NUM.get(priority, 3),
            "tags": tag_list,
        }
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            "https://ntfy.sh",
            data=data,
            method="POST",
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            resp.read()
        log.info("ntfy-уведомление отправлено (%s)", disp_title)
    except Exception as e:
        log.warning("notify_ntfy failed: %s", e)


def _configure_clipboard_signatures() -> None:
    """Выставляет argtypes/restype для Win32 API.

    КРИТИЧНО на 64-битной Windows: без явных restype ctypes считает,
    что функция возвращает C `int` (32 бита), и для функций возвращающих
    HANDLE/HGLOBAL/LPVOID (которые на x64 имеют ширину 64 бита) ctypes
    обрезает верхние 32 бита. Указатель превращается в мусор/0 →
    memmove(ptr, …) → access violation."""
    if getattr(_configure_clipboard_signatures, "_done", False):
        return
    u32 = ctypes.windll.user32
    k32 = ctypes.windll.kernel32
    u32.OpenClipboard.argtypes = [ctypes.c_void_p]
    u32.OpenClipboard.restype = ctypes.c_int
    u32.EmptyClipboard.restype = ctypes.c_int
    u32.CloseClipboard.restype = ctypes.c_int
    u32.SetClipboardData.argtypes = [ctypes.c_uint, ctypes.c_void_p]
    u32.SetClipboardData.restype = ctypes.c_void_p
    u32.GetClipboardData.argtypes = [ctypes.c_uint]
    u32.GetClipboardData.restype = ctypes.c_void_p
    k32.GlobalAlloc.argtypes = [ctypes.c_uint, ctypes.c_size_t]
    k32.GlobalAlloc.restype = ctypes.c_void_p
    k32.GlobalLock.argtypes = [ctypes.c_void_p]
    k32.GlobalLock.restype = ctypes.c_void_p
    k32.GlobalUnlock.argtypes = [ctypes.c_void_p]
    k32.GlobalUnlock.restype = ctypes.c_int
    k32.GlobalFree.argtypes = [ctypes.c_void_p]
    k32.GlobalFree.restype = ctypes.c_void_p
    _configure_clipboard_signatures._done = True


def _set_clipboard_win32(text: str) -> bool:
    """Кладёт текст в Windows clipboard через ctypes — без spawn'а PowerShell
    (раньше дочерний процесс мог стащить фокус с LS, и Ctrl+V улетал не туда)."""
    _configure_clipboard_signatures()
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    GMEM_MOVEABLE = 0x0002
    CF_UNICODETEXT = 13
    buf = (text + "\0").encode("utf-16-le")
    # 5 попыток — clipboard может быть кем-то открыт (антивирус, другой процесс)
    for _ in range(5):
        if user32.OpenClipboard(0):
            break
        time.sleep(0.1)
    else:
        return False
    h = None
    try:
        user32.EmptyClipboard()
        h = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(buf))
        if not h:
            return False
        ptr = kernel32.GlobalLock(h)
        if not ptr:
            kernel32.GlobalFree(h)
            return False
        ctypes.memmove(ptr, buf, len(buf))
        kernel32.GlobalUnlock(h)
        if not user32.SetClipboardData(CF_UNICODETEXT, h):
            kernel32.GlobalFree(h)
            return False
        # после успешного SetClipboardData владелец памяти — система,
        # GlobalFree больше не зовём.
        h = None
        return True
    finally:
        user32.CloseClipboard()


def _get_clipboard_win32() -> str | None:
    """Читает текст из clipboard для проверки, что Set прошёл."""
    _configure_clipboard_signatures()
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    CF_UNICODETEXT = 13
    if not user32.OpenClipboard(0):
        return None
    try:
        h = user32.GetClipboardData(CF_UNICODETEXT)
        if not h:
            return None
        ptr = kernel32.GlobalLock(h)
        if not ptr:
            return None
        try:
            return ctypes.wstring_at(ptr)
        finally:
            kernel32.GlobalUnlock(h)
    finally:
        user32.CloseClipboard()


class _KEYBDINPUT(ctypes.Structure):
    _fields_ = [
        ("wVk", ctypes.c_ushort),
        ("wScan", ctypes.c_ushort),
        ("dwFlags", ctypes.c_ulong),
        ("time", ctypes.c_ulong),
        ("dwExtraInfo", ctypes.c_void_p),
    ]


class _MOUSEINPUT(ctypes.Structure):
    _fields_ = [
        ("dx", ctypes.c_long),
        ("dy", ctypes.c_long),
        ("mouseData", ctypes.c_ulong),
        ("dwFlags", ctypes.c_ulong),
        ("time", ctypes.c_ulong),
        ("dwExtraInfo", ctypes.c_void_p),
    ]


class _HARDWAREINPUT(ctypes.Structure):
    _fields_ = [
        ("uMsg", ctypes.c_ulong),
        ("wParamL", ctypes.c_ushort),
        ("wParamH", ctypes.c_ushort),
    ]


class _INPUT_UNION(ctypes.Union):
    _fields_ = [("mi", _MOUSEINPUT), ("ki", _KEYBDINPUT), ("hi", _HARDWAREINPUT)]


class _INPUT(ctypes.Structure):
    _fields_ = [("type", ctypes.c_ulong), ("u", _INPUT_UNION)]


def _send_unicode_to_focused(text: str) -> bool:
    """Шлёт строку через SendInput с KEYEVENTF_UNICODE — каждый символ
    как Unicode-код в текущее focused окно, минуя клавиатурную раскладку
    И минуя Ctrl+V (т.е. нет проблем с тем, что shortcut не докатывается
    до Electron-инпута). Работает на любой language pack Windows.

    ВАЖНО: символы летят в то окно, где сейчас keyboard focus. Перед
    вызовом надо убедиться, что нужный input в фокусе (клик + sleep).
    Возвращает True если SendInput отправил все коды, False если ОС
    отвергла часть."""
    if sys.platform != "win32":
        return False
    user32 = ctypes.windll.user32
    user32.SendInput.argtypes = [ctypes.c_uint, ctypes.POINTER(_INPUT), ctypes.c_int]
    user32.SendInput.restype = ctypes.c_uint

    INPUT_KEYBOARD = 1
    KEYEVENTF_KEYUP = 0x0002
    KEYEVENTF_UNICODE = 0x0004

    # Каждый символ — два события (KEYDOWN + KEYUP), wScan = код-поинт.
    # BMP-символы (≤ 0xFFFF) шлём как один scan. Если бы были эмодзи
    # (> 0xFFFF) — нужно было бы surrogate pair; в email/password их нет.
    events: list[_INPUT] = []
    for ch in text:
        code = ord(ch)
        if code > 0xFFFF:
            # surrogate pair
            code -= 0x10000
            high = 0xD800 | (code >> 10)
            low = 0xDC00 | (code & 0x3FF)
            chars = [high, low]
        else:
            chars = [code]
        for scan in chars:
            for flags in (KEYEVENTF_UNICODE, KEYEVENTF_UNICODE | KEYEVENTF_KEYUP):
                inp = _INPUT()
                inp.type = INPUT_KEYBOARD
                inp.u.ki.wVk = 0
                inp.u.ki.wScan = scan
                inp.u.ki.dwFlags = flags
                inp.u.ki.time = 0
                inp.u.ki.dwExtraInfo = None
                events.append(inp)

    arr = (_INPUT * len(events))(*events)
    sent = user32.SendInput(len(events), arr, ctypes.sizeof(_INPUT))
    if sent != len(events):
        log.warning("SendInput отправил %d из %d событий", sent, len(events))
        return False
    return True


def _type_via_clipboard(text: str) -> None:
    """Вводит текст в текущий focused input.

    Главный путь — SendInput с KEYEVENTF_UNICODE: символы летят как
    Unicode-коды напрямую, нет Ctrl+V (не надо, чтобы шорткат докатился
    до Electron-input'а), нет зависимости от клавиатурной раскладки.

    Параллельно кладём текст и в clipboard — как safety net: если
    SendInput тоже промахнулся (поле не в фокусе), юзер сможет
    руками кликнуть в input и Ctrl+V. Данные уже в буфере."""
    if sys.platform != "win32":
        pyautogui.typewrite(text, interval=0.03)
        return
    # safety net: clipboard
    _set_clipboard_win32(text)
    # главный путь: SendInput Unicode
    if _send_unicode_to_focused(text):
        time.sleep(0.3)
        return
    # последний шанс: typewrite (зависит от раскладки, но лучше чем ничего)
    log.warning("SendInput Unicode failed → fallback typewrite")
    pyautogui.typewrite(text, interval=0.03)
    time.sleep(0.2)


class _RECT(ctypes.Structure):
    _fields_ = [("left", ctypes.c_long), ("top", ctypes.c_long),
                ("right", ctypes.c_long), ("bottom", ctypes.c_long)]


def _find_window_by_title_substring(substring: str) -> int:
    """Возвращает hwnd видимого top-level окна, заголовок которого содержит подстроку."""
    import ctypes.wintypes
    found = [0]
    sub_lower = substring.lower()

    @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)
    def _cb(hwnd: int, _: int) -> bool:
        if not ctypes.windll.user32.IsWindowVisible(hwnd):
            return True
        buf = ctypes.create_unicode_buffer(512)
        ctypes.windll.user32.GetWindowTextW(hwnd, buf, 512)
        if buf.value and sub_lower in buf.value.lower():
            found[0] = hwnd
            return False
        return True

    ctypes.windll.user32.EnumWindows(_cb, 0)
    return found[0]


def _find_window_by_any_title(substrings: list[str]) -> tuple[int, str]:
    """Возвращает (hwnd, matched_title) первого видимого top-level окна,
    заголовок которого содержит любую из подстрок. Регистронезависимо."""
    import ctypes.wintypes
    found: list[tuple[int, str]] = []
    subs_lower = [s.lower() for s in substrings]

    @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)
    def _cb(hwnd: int, _: int) -> bool:
        if not ctypes.windll.user32.IsWindowVisible(hwnd):
            return True
        buf = ctypes.create_unicode_buffer(512)
        ctypes.windll.user32.GetWindowTextW(hwnd, buf, 512)
        title = buf.value
        if not title:
            return True
        tl = title.lower()
        for sub in subs_lower:
            if sub in tl:
                found.append((hwnd, title))
                return False
        return True

    ctypes.windll.user32.EnumWindows(_cb, 0)
    return found[0] if found else (0, "")


def _log_visible_titles() -> None:
    """Логирует все видимые окна с непустым заголовком — для отладки."""
    import ctypes.wintypes
    titles: list[str] = []

    @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)
    def _cb(hwnd: int, _: int) -> bool:
        if not ctypes.windll.user32.IsWindowVisible(hwnd):
            return True
        buf = ctypes.create_unicode_buffer(512)
        ctypes.windll.user32.GetWindowTextW(hwnd, buf, 512)
        if buf.value:
            titles.append(buf.value)
        return True

    ctypes.windll.user32.EnumWindows(_cb, 0)
    log.info("видимые окна (%d): %s", len(titles), titles)


# Кэш, чтобы не спамить лог при каждом тике цикла ожидания
_visible_titles_logged = False


def _click_allow_button_via_message(parent_hwnd: int) -> bool:
    """Найти кнопку 'Allow access' среди дочерних окон Firewall-попапа и
    кликнуть через SendMessage(BM_CLICK). Это синтезирует клик НЕ через
    очередь ввода, поэтому не зависит от:
      • того, какое окно сейчас в фокусе (Alt+A улетал в LS-окно),
      • раскладки клавиатуры (русская/китайская тоже сработают),
      • DPI / RDP-скейлинга (без координат вообще).
    BM_CLICK = 0x00F5 — стандартное сообщение Windows для эмуляции клика
    по button-control'у. Адресуется конкретному child-hwnd, не глобально."""
    if sys.platform != "win32":
        return False

    user32 = ctypes.windll.user32
    user32.GetWindowTextW.argtypes = [ctypes.c_void_p, ctypes.c_wchar_p, ctypes.c_int]
    user32.GetWindowTextW.restype = ctypes.c_int
    user32.GetWindowTextLengthW.argtypes = [ctypes.c_void_p]
    user32.GetWindowTextLengthW.restype = ctypes.c_int
    user32.SendMessageW.argtypes = [ctypes.c_void_p, ctypes.c_uint, ctypes.c_void_p, ctypes.c_void_p]
    user32.SendMessageW.restype = ctypes.c_void_p
    user32.EnumChildWindows.argtypes = [ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p]

    candidates = ("allow access", "разрешить доступ")
    found = [None]

    ENUM_PROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)

    def _cb(hwnd, _lparam):
        length = user32.GetWindowTextLengthW(hwnd)
        if length <= 0:
            return True
        buf = ctypes.create_unicode_buffer(length + 1)
        user32.GetWindowTextW(hwnd, buf, length + 1)
        text = (buf.value or "").lower()
        for c in candidates:
            if c in text:
                found[0] = hwnd
                return False  # stop enum
        return True

    user32.EnumChildWindows(parent_hwnd, ENUM_PROC(_cb), 0)

    if found[0] is None:
        return False

    BM_CLICK = 0x00F5
    user32.SendMessageW(found[0], BM_CLICK, 0, 0)
    log.info("Firewall: BM_CLICK → Allow button hwnd=%s", found[0])
    return True


def _dismiss_firewall_alert() -> bool:
    """При первом запуске LS Windows может показать Defender Firewall Alert.
    Сначала пытается найти кнопку 'Allow access' по шаблону, иначе ищет окно
    по заголовку и кликает по координатам + Alt+A. Возвращает True если
    что-то закрыли."""
    global _visible_titles_logged
    if sys.platform != "win32":
        return False

    # 1) Шаблон allow_access.png — самый надёжный способ, не зависит от локали.
    # Заворачиваем в try, потому что ImageGrab.grab() падает с 'screen grab
    # failed' если RDP-сессия отключена. В этом случае проваливаемся на
    # BM_CLICK (он использует Win32 SendMessage, screen grab не нужен).
    try:
        if _click_allow_access_template():
            return True
    except OSError as e:
        log.warning("template-match для firewall упал (%s) — fallback на BM_CLICK", e)

    # 2) Заголовок зависит от локали Windows — пробуем все известные варианты
    titles = [
        "Windows Security Alert",
        "Windows Defender Firewall",
        "Брандмауэр",
        "Безопасность Windows",
        "Оповещение системы безопасности",
    ]
    hwnd, matched = _find_window_by_any_title(titles)
    if not hwnd:
        if not _visible_titles_logged:
            _log_visible_titles()
            _visible_titles_logged = True
        return False

    log.info("Firewall Alert hwnd=%d title=%r", hwnd, matched)
    user32 = ctypes.windll.user32

    # 3) BM_CLICK через SendMessage — самый надёжный способ, работает
    # независимо от фокуса и раскладки. На Win11 24H2/Server 2025 Alt+A
    # часто улетал в LS-окно (которое перехватывало фокус), а BM_CLICK
    # адресуется кнопке напрямую и не идёт через input queue.
    if _click_allow_button_via_message(hwnd):
        time.sleep(0.6)
        if not user32.IsWindow(hwnd):
            log.info("Firewall popup закрыт через BM_CLICK")
            return True
        log.warning("BM_CLICK отправлен, но окно ещё живо — fallback на Alt+A")

    def _force_foreground(target_hwnd: int) -> bool:
        """Возвращает True если окно ещё живо и попытка поднять отработала.
        Если окно закрылось (например, мы успешно его дисмиссили) — False,
        и дальше отправлять keystroke смысла нет."""
        if not user32.IsWindow(target_hwnd):
            return False
        user32.ShowWindow(target_hwnd, 9)   # SW_RESTORE
        user32.BringWindowToTop(target_hwnd)
        user32.SetForegroundWindow(target_hwnd)
        return True

    if not _force_foreground(hwnd):
        return False
    time.sleep(0.5)

    # 4) Alt+A — Windows-конвенция: 'A' подчёркнуто в 'Allow access', и
    # клавиша срабатывает на любом разрешении/DPI/языке UI (если только
    # язык не китайский с другими mnemonic). Fallback на случай если
    # BM_CLICK по какой-то причине не нашёл кнопку.
    log.info("Firewall dismiss: Alt+A")
    pyautogui.hotkey("alt", "a")
    time.sleep(1.0)

    if user32.IsWindow(hwnd):
        log.warning("Firewall popup ещё жив после Alt+A — возможно нестандартный диалог")
    else:
        log.info("Firewall popup закрыт")
    return True


def _match_template_in_region(
    name: str, confidence: float,
    left: int, top: int, right: int, bottom: int,
    scales: tuple[float, ...] = (0.75, 0.85, 0.95, 1.0, 1.1, 1.25, 1.5, 1.75, 2.0),
) -> tuple[int, int] | None:
    """Template-matching внутри прямоугольника экрана с перебором масштабов.
    cv2 не делает scale-invariant matching, поэтому если шаблон сохранён при
    одном DPI, а ищем при другом — нужно подставить размер. Возвращает
    абсолютные координаты центра лучшего матча, либо None."""
    tpl_path = TEMPLATES_DIR / f"{name}.png"
    tpl_orig = cv2.imread(str(tpl_path), cv2.IMREAD_COLOR)
    if tpl_orig is None:
        log.error("шаблон %s не читается", tpl_path)
        return None
    region_w = right - left
    region_h = bottom - top
    if region_w <= 0 or region_h <= 0:
        return None
    img = ImageGrab.grab(bbox=(left, top, right, bottom))
    region = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)

    best_val = 0.0
    best_loc = (0, 0)
    best_size = (0, 0)
    best_scale = 1.0
    orig_h, orig_w = tpl_orig.shape[:2]
    for s in scales:
        nw, nh = max(1, int(orig_w * s)), max(1, int(orig_h * s))
        if nw >= region_w or nh >= region_h:
            continue
        tpl = cv2.resize(tpl_orig, (nw, nh), interpolation=cv2.INTER_AREA if s < 1 else cv2.INTER_CUBIC)
        res = cv2.matchTemplate(region, tpl, cv2.TM_CCOEFF_NORMED)
        _, val, _, loc = cv2.minMaxLoc(res)
        if val > best_val:
            best_val = val
            best_loc = loc
            best_size = (nw, nh)
            best_scale = s

    log.info("match %s multiscale best=%.3f scale=%.2f @local(%d,%d) size=%dx%d",
             name, best_val, best_scale, best_loc[0], best_loc[1], *best_size)
    if best_val < confidence or best_size == (0, 0):
        return None
    return (left + best_loc[0] + best_size[0] // 2,
            top + best_loc[1] + best_size[1] // 2)


def _click_allow_access_template() -> bool:
    """Ищет шаблоны кнопок 'Allow' в Windows Defender Firewall Alert на всём
    экране. Версии Windows отличаются — кнопка может называться
    'Allow access' (со щитом UAC) или просто 'Allow' (Win11-style).
    Перебирает шаблоны по очереди:
      - allow_access   — Win10 стиль с щитом
      - allow_access2  — Win11 стиль без щита
    Возвращает True если что-то кликнули."""
    if sys.platform != "win32":
        return False

    import ctypes.wintypes
    screen_w = ctypes.windll.user32.GetSystemMetrics(0)
    screen_h = ctypes.windll.user32.GetSystemMetrics(1)

    for tpl_name in ("allow_access", "allow_access2"):
        if not (TEMPLATES_DIR / f"{tpl_name}.png").exists():
            continue
        # Понижаем порог до 0.70 СПЕЦИАЛЬНО для allow-кнопок: шаблоны 180×53
        # и 249×60 в multi-scale матчинге часто дают best ~0.65-0.72 из-за
        # DPI/UAC-shield render. 0.80 был слишком строгий, ловили миссы.
        pt = _match_template_in_region(
            tpl_name, 0.70,
            0, 0, screen_w, screen_h,
        )
        if pt is None:
            continue
        log.info("%s кнопка найдена @(%d,%d)", tpl_name, *pt)
        _record_click(pt[0], pt[1], tpl_name)
        pyautogui.click(pt[0], pt[1])
        time.sleep(1.0)
        return True
    return False


def _wait_for_firewall_alert(seconds: float, exit_after_close: bool = False) -> bool:
    """Активно следит `seconds` секунд за появлением Windows Firewall Alert.
    Жмёт 'Allow access' двумя способами: сначала шаблон allow_access.png,
    потом fallback на window-title + Alt+A. NEXT STEP в этот момент НЕ жмём.
    Если exit_after_close=True — выходит сразу после первого успешного
    закрытия (плюс 3с grace на случай повторного диалога).
    Возвращает True если закрыл хоть один alert."""
    log.info("ждём до %.0fс появления firewall alert (NEXT STEP пока НЕ жмём)…", seconds)
    deadline = time.time() + seconds
    closed_any = False
    grace_deadline: float | None = None
    while time.time() < deadline:
        if _click_allow_access_template():
            closed_any = True
            if exit_after_close:
                grace_deadline = time.time() + 3.0
            continue
        if _dismiss_firewall_alert():
            closed_any = True
            if exit_after_close:
                grace_deadline = time.time() + 3.0
            continue
        if grace_deadline is not None and time.time() > grace_deadline:
            log.info("firewall alert закрыт, выходим из ожидания пораньше")
            return True
        time.sleep(0.5)
    if closed_any:
        log.info("firewall alert закрыт")
    else:
        log.info("firewall alert не появился за отведённое время")
    return closed_any


def _find_ls_window() -> int:
    """Находит главное окно Linken Sphere (заголовок 'Linken Sphere',
    крупный размер, на экране). Wizard первого запуска показывается ВНУТРИ
    этого же окна, отдельного top-level окна у мастера нет."""
    if sys.platform != "win32":
        return 0
    import ctypes.wintypes
    candidates: list[int] = []

    @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)
    def _cb(hwnd: int, _: int) -> bool:
        if not ctypes.windll.user32.IsWindowVisible(hwnd):
            return True
        buf = ctypes.create_unicode_buffer(256)
        ctypes.windll.user32.GetWindowTextW(hwnd, buf, 256)
        title = buf.value
        if title and "linken" in title.lower() and "sphere" in title.lower():
            candidates.append(hwnd)
        return True

    ctypes.windll.user32.EnumWindows(_cb, 0)
    for hwnd in candidates:
        rect = _RECT()
        ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
        w = rect.right - rect.left
        h = rect.bottom - rect.top
        if w >= 600 and h >= 400 and rect.left > -1000 and rect.top > -1000:
            return hwnd
    return 0


# Состояние для close_x: ищем крестик ✕ только в течение этого окна
# после успешного клика на SKIP. Иначе шаблон ложно матчит close-кнопку
# самой LS на повторных запусках, когда попапа со скипом уже нет.
_skip_clicked_at: float = 0.0
_CLOSE_X_GRACE_AFTER_SKIP = 30.0


def _dismiss_customize_wizard_step() -> bool:
    """При первом запуске LS показывает мастер настройки ВНУТРИ окна Linken Sphere
    (отдельного top-level окна у мастера нет). Ищем кнопки внутри окна LS —
    если есть, кликаем. Шаблоны по очереди:
      - next_step       — первые две страницы wizard
      - get_started     — последняя страница wizard ('Get Started >')
      - get_started2    — приветственный экран после логина
                          ('Welcome to Linken Sphere 2', другой фон)
      - skip            — иногда всплывающее окно с малозаметной кнопкой SKIP
      - close_x         — финальный мелкий крестик ✕ на следующем после skip окне;
                          ищется ТОЛЬКО в течение 30с после клика на skip,
                          чтобы не зацепить close-кнопку самого LS
    Когда все исчезнут — функция вернёт False, поток пойдёт дальше."""
    global _skip_clicked_at
    if sys.platform != "win32":
        return False

    hwnd = _find_ls_window()
    if not hwnd:
        return False

    rect = _RECT()
    ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
    h = rect.bottom - rect.top
    lower_half_top = rect.top + h // 2

    # (имя шаблона, top границы поиска, confidence)
    candidates = [
        ("next_step",    lower_half_top, 0.80),
        ("get_started",  lower_half_top, 0.80),
        ("get_started2", lower_half_top, 0.80),
        ("skip",         rect.top,       0.85),
    ]
    # close_x активен только в окне 30с после успешного клика на skip
    if time.time() - _skip_clicked_at < _CLOSE_X_GRACE_AFTER_SKIP:
        candidates.append(("close_x", rect.top, 0.90))

    for tpl_name, search_top, conf in candidates:
        if not (TEMPLATES_DIR / f"{tpl_name}.png").exists():
            continue

        # close_x опасен: его шаблон визуально близок к close-кнопке самой LS
        # в правом верхнем углу окна. Исключаем верхнюю title-bar полосу и
        # правые 100px (где сидит min/max/close самого LS).
        if tpl_name == "close_x":
            search_left = rect.left
            search_top_eff = rect.top + 50
            search_right = rect.right - 100
        else:
            search_left = rect.left
            search_top_eff = search_top
            search_right = rect.right

        pt = _match_template_in_region(
            tpl_name, conf,
            search_left, search_top_eff, search_right, rect.bottom,
        )
        if pt is not None:
            ctypes.windll.user32.SetForegroundWindow(hwnd)
            time.sleep(0.3)
            log.info("wizard: %s найден в окне LS @(%d,%d)", tpl_name.upper(), *pt)
            _record_click(pt[0], pt[1], tpl_name)
            pyautogui.click(pt[0], pt[1])
            if tpl_name == "skip":
                _skip_clicked_at = time.time()
            time.sleep(1.5)
            return True
    return False


def _find_auth_window() -> int:
    """Ищет окно Linken Sphere 2 с формой входа. Разворачивает если свёрнуто.
    Пропускает мелкие/служебные окна Electron с тем же заголовком."""
    import ctypes.wintypes

    candidates: list[int] = []
    all_titles: list[str] = []

    @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)
    def _cb(hwnd: int, _: int) -> bool:
        if not ctypes.windll.user32.IsWindowVisible(hwnd):
            return True
        buf = ctypes.create_unicode_buffer(256)
        ctypes.windll.user32.GetWindowTextW(hwnd, buf, 256)
        title = buf.value
        if not title:
            return True
        all_titles.append(title)
        tl = title.lower()
        if "authentication" in tl or ("linken" in tl and "sphere" in tl):
            candidates.append(hwnd)
        return True

    ctypes.windll.user32.EnumWindows(_cb, 0)
    log.info("видимые окна: %s", all_titles)
    log.info("кандидатов на окно входа: %d", len(candidates))

    SW_RESTORE = 9
    for hwnd in candidates:
        if ctypes.windll.user32.IsIconic(hwnd):
            log.info("hwnd=%d свёрнуто, разворачиваю", hwnd)
            ctypes.windll.user32.ShowWindow(hwnd, SW_RESTORE)
            time.sleep(0.6)

        rect = _RECT()
        ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
        w = rect.right - rect.left
        h = rect.bottom - rect.top
        log.info("hwnd=%d %dx%d @ (%d,%d)", hwnd, w, h, rect.left, rect.top)

        # Окно формы логина должно быть крупным и на экране
        if w >= 600 and h >= 400 and rect.left > -1000 and rect.top > -1000:
            return hwnd

    if candidates:
        log.warning("найдены окна Linken Sphere, но ни одно не похоже на форму входа (мелкие/за экраном)")
    return 0


def login_if_needed(cfg: configparser.ConfigParser) -> None:
    """Если видна форма аутентификации — входим по данным из credentials.ini."""
    conf = cfg.getfloat("matching", "confidence")

    if find_template("three_dots", conf) is not None:
        log.info("login_if_needed: уже на главном экране")
        return

    if sys.platform != "win32":
        return

    hwnd = _find_auth_window()
    if not hwnd:
        log.info("login_if_needed: окно аутентификации не найдено — пропуск")
        return

    log.info("login_if_needed: обнаружен экран входа (hwnd=%d)", hwnd)

    # КРИТИЧНО: до клика по форме входа закрыть всё модальное, что может быть
    # поверх LS (Windows Defender Firewall Alert, мастер Customize). Иначе
    # клики по координатам логин-формы попадут в этот диалог, а не в LS.
    for _ in range(10):
        did = False
        if _dismiss_firewall_alert():
            did = True
        if _dismiss_customize_wizard_step():
            did = True
        if not did:
            break

    screenshot("login_screen", cfg.getboolean("logging", "screenshots"))

    rect = _RECT()
    ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
    win_w = rect.right - rect.left
    win_h = rect.bottom - rect.top
    cx = rect.left + win_w // 2

    if win_w < 600 or win_h < 400 or rect.left < -1000 or rect.top < -1000:
        log.error("окно %dx%d @ (%d,%d) — невалидное (за экраном/мелкое), отменяю логин",
                  win_w, win_h, rect.left, rect.top)
        return

    # Пропорции замерены по скриншоту: Email=43%, Password=52%, SIGN IN=68% от высоты
    email_y    = rect.top + int(win_h * 0.434)
    password_y = rect.top + int(win_h * 0.521)
    signin_y   = rect.top + int(win_h * 0.678)

    log.info("окно %dx%d @ (%d,%d)", win_w, win_h, rect.left, rect.top)
    log.info("email=(%d,%d) password=(%d,%d) signin=(%d,%d)",
             cx, email_y, cx, password_y, cx, signin_y)

    ctypes.windll.user32.SetForegroundWindow(hwnd)
    time.sleep(0.5)

    creds = load_credentials()

    _record_click(cx, email_y, "email")
    pyautogui.click(cx, email_y)
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "a")
    _type_via_clipboard(creds.get("account", "email"))

    _record_click(cx, password_y, "password")
    pyautogui.click(cx, password_y)
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "a")
    _type_via_clipboard(creds.get("account", "password"))

    screenshot("login_filled", cfg.getboolean("logging", "screenshots"))

    _record_click(cx, signin_y, "sign_in")
    pyautogui.click(cx, signin_y)
    log.info("нажат SIGN IN, жду главный экран…")
    time.sleep(5.0)  # сетевая сторона signin на слабом VPS отзывается с лагом

    # После SIGN IN ждём three_dots, но параллельно чистим всплывающие диалоги
    timeout = cfg.getfloat("startup", "launch_wait_seconds")
    deadline = time.time() + timeout
    while time.time() < deadline:
        if _dismiss_firewall_alert():
            continue
        if _dismiss_customize_wizard_step():
            continue
        if find_template("three_dots", conf) is not None:
            log.info("вход выполнен, главный экран готов")
            time.sleep(1.0)
            return
        time.sleep(0.5)
    raise TimeoutError(f"three_dots не появился за {timeout}s после SIGN IN")


def ensure_linken_sphere_running(cfg: configparser.ConfigParser) -> None:
    """
    Если главный экран Linken Sphere 2 (по three_dots.png) уже виден — ничего не делаем.
    Если уже на экране входа — тоже не перезапускаем (login_if_needed сделает вход).
    Иначе мягко завершаем старые процессы, запускаем через ShellExecuteW и ждём.
    """
    conf = cfg.getfloat("matching", "confidence")

    # Первые 3 секунды: ищем three_dots, но параллельно чистим firewall alert
    # и мастер первого запуска — они могут всплыть с задержкой после старта warmup.
    quick_deadline = time.time() + 3.0
    found_main = False
    while time.time() < quick_deadline:
        if _dismiss_firewall_alert():
            continue
        if _dismiss_customize_wizard_step():
            continue
        if find_template("three_dots", conf) is not None:
            found_main = True
            break
        time.sleep(0.4)
    if found_main:
        log.info("Linken Sphere 2 уже на главном экране")
        return

    # Уже показывает экран входа — не убиваем, просто дадим login_if_needed войти
    if sys.platform == "win32" and _find_auth_window():
        log.info("Linken Sphere 2 уже запущен (экран входа), не перезапускаем")
        return

    # Уже висит мастер первого запуска — пройдём его и продолжим
    if sys.platform == "win32" and _find_window_by_title_substring("Customize your experience"):
        log.info("Linken Sphere 2 уже запущен (мастер первого запуска), прокликиваю")
        deadline_wiz = time.time() + 60.0
        while time.time() < deadline_wiz:
            if _dismiss_firewall_alert():
                continue
            if _dismiss_customize_wizard_step():
                continue
            if find_template("three_dots", conf) is not None:
                log.info("после мастера — главный экран")
                return
            if _find_auth_window():
                log.info("после мастера — экран входа")
                return
            time.sleep(0.5)

    path = cfg.get("startup", "linken_sphere_path")
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"не найден путь к Linken Sphere 2: {path}. "
            "поправь startup.linken_sphere_path в config.ini"
        )

    exe_name = os.path.basename(path)

    # Мягкое закрытие (WM_CLOSE)
    r = subprocess.run(["taskkill", "/im", exe_name], capture_output=True, text=True)
    log.info("taskkill soft: stdout=%r stderr=%r", r.stdout.strip(), r.stderr.strip())
    time.sleep(3.0)

    # Принудительное завершение остатков
    r = subprocess.run(["taskkill", "/f", "/im", exe_name, "/t"], capture_output=True, text=True)
    log.info("taskkill force: stdout=%r stderr=%r", r.stdout.strip(), r.stderr.strip())
    time.sleep(3.0)

    exe_dir = os.path.dirname(path)
    log.info("запускаю через ShellExecuteW: %s (cwd=%s)", path, exe_dir)

    # ShellExecuteW = точно как двойной клик в Проводнике
    ret = ctypes.windll.shell32.ShellExecuteW(None, "open", path, None, exe_dir, 1)
    log.info("ShellExecuteW → %d (%s)", ret, "OK" if ret > 32 else "ОШИБКА")
    if ret <= 32:
        log.warning("ShellExecuteW вернул ошибку %d, запускаю через Popen", ret)
        subprocess.Popen(
            [path],
            cwd=exe_dir,
            creationflags=subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP,
        )

    log.info("жду инициализацию Electron (5s базовая пауза)…")
    time.sleep(5.0)

    # Первые 10 секунд после старта LS — следим ТОЛЬКО за firewall alert
    # (Windows Defender Firewall). NEXT STEP в это время НЕ жмём, чтобы
    # системный диалог успел всплыть и был обработан без перекрытия кликами.
    _wait_for_firewall_alert(10.0)

    timeout = cfg.getfloat("startup", "launch_wait_seconds")
    log.info("жду главный экран или экран входа (до %.0fs)…", timeout)
    deadline_ls = time.time() + timeout
    found_screen = None
    while time.time() < deadline_ls:
        # При первом запуске LS на чистой машине могут всплыть диалоги —
        # закрываем их прежде, чем смотреть на основные экраны.
        if _dismiss_firewall_alert():
            continue
        if _dismiss_customize_wizard_step():
            continue
        if find_template("three_dots", conf) is not None:
            found_screen = "main"
            break
        if sys.platform == "win32" and _find_auth_window():
            found_screen = "login"
            break
        time.sleep(0.5)
    if found_screen is None:
        raise TimeoutError("Linken Sphere не показал экран за отведённое время")
    log.info("Linken Sphere готов (экран: %s)", found_screen)
    time.sleep(1.0)


def _files_dir(cfg: configparser.ConfigParser) -> str:
    """Резолвит paths.files_dir: абсолютный путь возвращает как есть, относительный — относительно ROOT."""
    p = cfg.get("paths", "files_dir")
    if not os.path.isabs(p):
        p = str(ROOT / p)
    return p


def pick_random_file(cfg: configparser.ConfigParser) -> str:
    """Сэмплит ~100 случайных URL из [api] url_pool_file (40k_all_urls.txt)
    и материализует временный файл urls_generated/manual_<ts>.txt — он
    привязывается к UI warmup через Browse File. Тот же источник, что
    использует warmup_api.py, симметрия между ручным и автоматическим
    флоу."""
    pool_rel = cfg.get("api", "url_pool_file", fallback="urls/40k_all_urls.txt")
    pool_path = ROOT / pool_rel
    if not pool_path.exists():
        raise FileNotFoundError(f"URL-пул не найден: {pool_path}")
    pool: list[str] = []
    seen: set[str] = set()
    for line in pool_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        u = line.strip()
        if u and u not in seen:
            seen.add(u)
            pool.append(u)
    if not pool:
        raise RuntimeError(f"в {pool_path} нет ни одного URL")
    n_min = cfg.getint("api", "urls_per_run_min", fallback=95)
    n_max = cfg.getint("api", "urls_per_run_max", fallback=105)
    n = min(random.randint(n_min, n_max), len(pool))
    urls = random.sample(pool, n)

    out_dir = ROOT / "urls_generated"
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = time.strftime("%Y%m%d-%H%M%S")
    out_file = out_dir / f"manual_{ts}.txt"
    out_file.write_text("\n".join(urls) + "\n", encoding="utf-8")
    log.info("выбрано %d URL из пула %d → %s", n, len(pool), out_file)
    return str(out_file)


def handle_open_file_dialog(file_path: str, cfg: configparser.ConfigParser) -> None:
    """
    Нативный Windows-диалог открытия файла.
    Самый надёжный способ — вставить полный путь в поле "Имя файла" и нажать Enter.
    """
    time.sleep(3.0)  # нативный диалог Windows на 2-ядерном VPS отрисовывается дольше
    screenshot("06_open_dialog", cfg.getboolean("logging", "screenshots"))
    # focus на поле "File name" — стандартно открыто по умолчанию
    pyautogui.hotkey("alt", "n")  # Alt+N → File name (en)
    time.sleep(0.5)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("delete")
    # _type_via_clipboard использует SendInput Unicode — на той же скорости,
    # что и pyautogui.typewrite interval=0.01, проглатывал буквы из пути,
    # и Windows ругался «The file name is not valid».
    _type_via_clipboard(file_path)
    time.sleep(0.3)
    pyautogui.press("enter")
    time.sleep(cfg.getfloat("matching", "step_delay"))


def _generate_session_name() -> str:
    """CL-XXXXXXXX, 8 цифр. 10^8 возможных значений — для пары сотен VPS
    вероятность пересечения <0.001%."""
    return f"CL-{random.randint(10_000_000, 99_999_999)}"


def load_session_name() -> str:
    """Имя сессии этой машины. Источники, по приоритету:
      1. .session_name (gitignored) — главный источник, генерируется автоматом
         на первом запуске и больше не меняется.
      2. .session_imported (старый формат «только имя») — миграция со старых
         инсталляций, где имя жило там.
      3. credentials.ini [session] name — миграция со ещё более старого
         формата (когда юзер вписывал имя руками).
      4. Если ничего нет — генерим новое CL-XXXXXXXX и пишем в .session_name.
    Всегда возвращает непустую строку."""
    if SESSION_NAME_FILE.exists():
        name = SESSION_NAME_FILE.read_text(encoding="utf-8").strip()
        if name:
            return name
    if SESSION_IMPORTED_FLAG.exists():
        legacy = SESSION_IMPORTED_FLAG.read_text(encoding="utf-8").strip()
        if legacy and "\t" not in legacy:
            log.info("миграция: имя сессии %r из .session_imported → .session_name", legacy)
            SESSION_NAME_FILE.write_text(legacy, encoding="utf-8")
            return legacy
    try:
        creds = load_credentials()
        if creds.has_section("session"):
            legacy = creds.get("session", "name", fallback="").strip()
            if legacy and legacy.lower() not in ("session name", "your session name", "yourname"):
                log.info("миграция: имя сессии %r из credentials.ini → .session_name", legacy)
                SESSION_NAME_FILE.write_text(legacy, encoding="utf-8")
                return legacy
    except FileNotFoundError:
        pass
    name = _generate_session_name()
    SESSION_NAME_FILE.write_text(name, encoding="utf-8")
    log.info("сгенерировано новое имя сессии: %s", name)
    return name


def prepare_session_xlsx(session_name: str) -> Path:
    """Если session_imports/<name>.xlsx уже есть — возвращаем путь.
    Иначе клонируем session_imports/_template.xlsx, ставим A3 = session_name,
    сохраняем под нужным именем. Возвращаем путь к готовому файлу."""
    target = SESSION_IMPORTS_DIR / f"{session_name}.xlsx"
    if target.exists():
        return target
    template = SESSION_IMPORTS_DIR / "_template.xlsx"
    if not template.exists():
        raise FileNotFoundError(
            f"не найден шаблон импорта сессии: {template}. "
            f"Должен лежать в репозитории."
        )
    import openpyxl  # ленивый импорт — нужен только при первой установке
    wb = openpyxl.load_workbook(template)
    ws = wb[wb.sheetnames[0]]
    ws["A3"] = session_name
    wb.save(target)
    log.info("создан xlsx сессии: %s (A3=%s)", target, session_name)
    return target


def _click_import_browse_file(cfg: configparser.ConfigParser) -> None:
    """В окне Mass creation две одинаковые кнопки BROWSE FILE: левая — cookies
    (TXT/JSON), правая — XLSX/CSV. Нам нужна ПРАВАЯ, поэтому ищем шаблон только
    в правой половине экрана."""
    conf = cfg.getfloat("matching", "confidence")
    timeout = cfg.getfloat("matching", "wait_seconds")
    sw = ctypes.windll.user32.GetSystemMetrics(0)
    sh = ctypes.windll.user32.GetSystemMetrics(1)
    deadline = time.time() + timeout
    while time.time() < deadline:
        pt = _match_template_in_region("browse_file", conf, sw // 2, 0, sw, sh)
        if pt is not None:
            log.info("BROWSE FILE (правая) @(%d,%d)", *pt)
            _record_click(pt[0], pt[1], "browse_file")
            pyautogui.click(pt[0], pt[1])
            time.sleep(cfg.getfloat("matching", "step_delay"))
            return
        time.sleep(0.5)
    raise TimeoutError("BROWSE FILE (правая, XLSX) не найдена в правой половине экрана")


def _check_api_port_alive(base_url: str) -> bool:
    """Пинг локального LS API — пробуем GET /sessions. 4xx/5xx ок (порт жив),
    сеть/таймаут → порт не открыт."""
    import urllib.request, urllib.error
    try:
        req = urllib.request.Request(f"{base_url.rstrip('/')}/sessions", method="GET")
        with urllib.request.urlopen(req, timeout=3) as resp:
            resp.read()
        return True
    except urllib.error.HTTPError:
        return True
    except (urllib.error.URLError, OSError):
        return False


def _parse_api_port_from_config(cfg: configparser.ConfigParser) -> tuple[str, int]:
    """Тянет base_url из config.ini [api] и вытаскивает оттуда порт.
    Возвращает (base_url, port)."""
    base_url = cfg.get("api", "base_url", fallback="http://127.0.0.1:36555").rstrip("/")
    # base_url вида http://host:port — порт после последнего двоеточия
    try:
        port = int(base_url.rsplit(":", 1)[1])
    except (IndexError, ValueError):
        port = 36555
    return base_url, port


def activate_api_port_if_needed(cfg: configparser.ConfigParser) -> None:
    """Одноразово на машине: открывает Settings → прокручивает вниз →
    вписывает API-порт → Enter → Esc. После успешного пинга API пишет
    флаг .api_activated, дальше пропускает.

    Защитный dismiss-loop в начале — на случай, если после login_if_needed
    всплыл get_started/get_started2/skip/firewall, который ещё не успели
    закрыть."""
    if API_ACTIVATED_FLAG.exists():
        log.info("API-порт уже активирован ранее — пропуск")
        return

    base_url, port = _parse_api_port_from_config(cfg)

    # Быстрая проверка: вдруг порт уже активирован руками — не надо лезть в UI.
    if _check_api_port_alive(base_url):
        log.info("API на %s уже отвечает — ставим флаг без UI-активации", base_url)
        API_ACTIVATED_FLAG.write_text(str(port), encoding="utf-8")
        return

    log.info("активирую API-порт %d через UI", port)
    shots = cfg.getboolean("logging", "screenshots")

    # 1. Закрыть всё что могло всплыть после логина (wizard / firewall).
    for _ in range(10):
        did = False
        if _dismiss_firewall_alert():
            did = True
        if _dismiss_customize_wizard_step():
            did = True
        if not did:
            break

    # 2. Клик на settings_gear (маленькая шестерёнка слева вверху LS).
    #    Шаблон 32×28 — слишком мелкий, чтобы безопасно искать по всему
    #    экрану (любая чёрная иконка в браузере/трее может сматчиться).
    #    Ограничиваем регион верхним-левым углом окна LS.
    ls_hwnd = _find_ls_window()
    conf = cfg.getfloat("matching", "confidence")
    timeout = cfg.getfloat("matching", "wait_seconds")
    pt = None
    if ls_hwnd:
        rect = _RECT()
        ctypes.windll.user32.GetWindowRect(ls_hwnd, ctypes.byref(rect))
        # шестерёнка живёт в шапке LS — слева, в первых ~250×120 пикселях
        deadline = time.time() + timeout
        while time.time() < deadline:
            pt = _match_template_in_region(
                "settings_gear", conf,
                rect.left, rect.top, rect.left + 260, rect.top + 130,
            )
            if pt is not None:
                break
            time.sleep(0.5)
    if pt is None:
        # fallback — попробуем full-screen матч (если LS hwnd не найден или
        # шапка нестандартная); confidence уже 0.80, риск false positive есть,
        # но это последняя надежда.
        log.warning("settings_gear не нашли в шапке LS, пробуем по всему экрану")
        click("settings_gear", cfg)
    else:
        log.info("settings_gear @(%d,%d)", *pt)
        _record_click(pt[0], pt[1], "settings_gear")
        pyautogui.click(pt[0], pt[1])
        time.sleep(cfg.getfloat("matching", "step_delay"))
    screenshot("A1_preferences_open", shots)

    # 3. Скроллим к самому низу страницы Preferences — поле Api port внизу.
    #    Сначала клик в центр области Preferences, чтобы скролл-фокус
    #    точно был на содержимом, а не на каком-нибудь баннере.
    sw = ctypes.windll.user32.GetSystemMetrics(0)
    sh = ctypes.windll.user32.GetSystemMetrics(1)
    pyautogui.click(sw // 2, sh // 2)
    time.sleep(0.5)
    # Несколько Ctrl+End — Electron иногда обрабатывает только первый.
    for _ in range(3):
        pyautogui.hotkey("ctrl", "end")
        time.sleep(0.6)
    # Дополнительно PgDn много раз — на случай если Ctrl+End не сработал.
    for _ in range(15):
        pyautogui.press("pagedown")
        time.sleep(0.15)
    time.sleep(1.0)
    screenshot("A2_scrolled_to_api_port", shots)

    # 4. Найти api_port_field (строка «Api port» + input справа).
    #    Кликаем в правую часть найденного бокса — там input.
    box = _find_template_box("api_port_field", cfg)
    field_left, field_top, field_w, field_h = box
    click_x = field_left + int(field_w * 0.80)   # ~80% по ширине — гарантированно в input
    click_y = field_top + field_h // 2
    log.info("Api port input @(%d,%d) box=%dx%d", click_x, click_y, field_w, field_h)
    _record_click(click_x, click_y, "api_port_input")
    pyautogui.click(click_x, click_y)
    time.sleep(0.5)

    # 5. Очистить (вдруг там уже что-то по дефолту) и вписать порт.
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.2)
    pyautogui.press("delete")
    time.sleep(0.2)
    _type_via_clipboard(str(port))
    time.sleep(0.5)
    screenshot("A3_port_typed", shots)

    # 6. Enter — подтвердить значение.
    pyautogui.press("enter")
    time.sleep(1.0)

    # 7. Esc — выйти из Preferences обратно на список сессий.
    pyautogui.press("escape")
    time.sleep(1.5)
    screenshot("A4_back_to_main", shots)

    # 8. Сanity: API теперь должен отвечать. Иначе флаг не ставим.
    if not _check_api_port_alive(base_url):
        raise RuntimeError(
            f"API-порт {port} вписан, Enter нажат, Esc нажат — но {base_url} "
            f"не отвечает. Скорее всего LS не сохранила порт. Проверь "
            f"скриншоты A1-A4 в screenshots/."
        )

    API_ACTIVATED_FLAG.write_text(str(port), encoding="utf-8")
    log.info("API-порт %d активирован, флаг записан", port)


def import_session_if_needed(cfg: configparser.ConfigParser, session_name: str) -> None:
    """Первый запуск на машине: импортит сессию из session_imports/<name>.xlsx
    через MULTIPLE → BROWSE FILE (правая) → диалог → IMPORT. Дальше пропускается."""
    if SESSION_IMPORTED_FLAG.exists():
        prev = SESSION_IMPORTED_FLAG.read_text(encoding="utf-8").strip()
        if prev == session_name:
            log.info("сессия %r уже импортирована ранее — пропуск импорта", session_name)
            return
        log.info("в флаге другое имя (%r != %r) — импортирую заново", prev, session_name)

    xlsx = prepare_session_xlsx(session_name)

    shots = cfg.getboolean("logging", "screenshots")
    log.info("импорт сессии %r из %s", session_name, xlsx)

    # подстраховка: закрыть firewall, если всплыл перед импортом
    _dismiss_firewall_alert()

    # 1. MULTIPLE → окно Mass creation
    click("multiple_button", cfg)
    screenshot("D1_mass_creation", shots)

    # 2. BROWSE FILE (правая, XLSX)
    _click_import_browse_file(cfg)

    # 3. нативный диалог: вписать путь и Enter
    handle_open_file_dialog(str(xlsx), cfg)
    screenshot("D2_file_chosen", shots)

    # 4. IMPORT
    click("import_button", cfg)
    log.info("нажат IMPORT, жду создания сессии (10с) с попутным дисмиссом firewall…")
    # импорт может дёрнуть сеть → возможен firewall alert; ждём 10с и чистим его
    _wait_for_firewall_alert(10.0)
    screenshot("D3_after_import", shots)

    SESSION_IMPORTED_FLAG.write_text(session_name, encoding="utf-8")
    log.info("сессия импортирована, флаг записан")


def search_session(cfg: configparser.ConfigParser, session_name: str) -> None:
    """Фокусирует поиск Сферы (Ctrl+F) и вписывает имя сессии — в списке
    остаётся одна строка, three_dots на ней будет однозначным."""
    log.info("поиск сессии по имени: %r", session_name)
    pyautogui.hotkey("ctrl", "f")
    time.sleep(1.2)  # поле поиска появляется не мгновенно на слабом UI
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("delete")
    _type_via_clipboard(session_name)
    time.sleep(2.5)  # фильтрация списка на 2-ядерном VPS заметно медленнее
    screenshot("D4_searched", cfg.getboolean("logging", "screenshots"))


def run() -> int:
    cfg = load_config()
    shots = cfg.getboolean("logging", "screenshots")
    log.info("=" * 60)
    log.info("Linken Sphere warm-up: старт сценария")

    try:
        file_to_attach = pick_random_file(cfg)

        # 0. Запустить Linken Sphere 2, если ещё не запущен
        ensure_linken_sphere_running(cfg)

        # 0.5 Войти, если показан экран аутентификации
        login_if_needed(cfg)

        screenshot("00_initial", shots)

        # 0.6 Активировать API-порт LS (Settings → Network → Api port).
        #     Одноразово на машине. После — все прогревы через HTTP API,
        #     UI больше не дёргаем. activate_api_port_if_needed сам
        #     дисмиссит wizard/firewall если что-то всплыло после логина.
        activate_api_port_if_needed(cfg)

        # 0.7 Сессия машины: уникальное имя CL-XXXXXXXX генерится один раз
        #     и хранится в .session_name. Первый запуск — клонируем шаблон
        #     xlsx с этим именем + импорт в LS, затем поиск по имени, чтобы
        #     three_dots был однозначным.
        session_name = load_session_name()
        import_session_if_needed(cfg, session_name)
        search_session(cfg, session_name)
        # подстраховка перед three_dots: закрыть firewall, если всплыл
        _dismiss_firewall_alert()

        # 1. меню "три точки"
        click("three_dots", cfg)
        screenshot("01_after_three_dots", shots)

        # 2. пункт Warm up
        click("warm_up_menu", cfg)
        screenshot("02_warmup_window", shots)

        # 3. viewing depth (stepper)
        set_stepper(
            "viewing_depth_field",
            cfg.getint("warmup", "viewing_depth"),
            cfg.getint("warmup", "viewing_depth_min"),
            cfg,
        )
        screenshot("03_viewing_depth_set", shots)

        # 4. time per url (stepper)
        set_stepper(
            "time_per_url_field",
            cfg.getint("warmup", "time_per_url"),
            cfg.getint("warmup", "time_per_url_min"),
            cfg,
        )
        screenshot("04_time_per_url_set", shots)

        # 5. снять галку — toggle слева в шаблоне, не по центру
        if cfg.getboolean("warmup", "uncheck_use_most_popular"):
            click_at_offset(
                "use_most_popular_checkbox",
                cfg.getfloat("checkbox_offset", "toggle_x"),
                cfg.getfloat("checkbox_offset", "toggle_y"),
                cfg,
            )
            screenshot("05_unchecked", shots)

        # 6. Browse file → диалог → ввести путь → Enter
        click("browse_file_button", cfg)
        handle_open_file_dialog(file_to_attach, cfg)
        screenshot("07_after_browse", shots)

        # 6.5. Удалить первые N строк из URL-списка (если задано)
        remove_n = cfg.getint("warmup", "remove_first_n_lines", fallback=0)
        if remove_n > 0:
            remove_first_lines_from_list(remove_n, cfg)
            screenshot("07b_lines_removed", shots)

        # 7. START
        click("start_button", cfg)
        screenshot("08_started", shots)

        # 7.5. После START LS греет URL'ы ~40 минут — Windows может в любой
        # момент показать firewall alert. Запускаем фоновый watcher-процесс,
        # который дисмиссит firewall каждые 15с. Ждём примерно столько же,
        # сколько длится прогрев (100 URL × ~25с = ~42 мин, берём с запасом).
        warmup_seconds = cfg.getint("warmup", "time_per_url", fallback=7) * 100 * 4
        log.info("watcher firewall на %dс (длительность прогрева + запас)", warmup_seconds)
        fw_proc = subprocess.Popen(
            [sys.executable, str(ROOT / "_firewall_watcher.py"), "15"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        try:
            time.sleep(warmup_seconds)
        finally:
            fw_proc.terminate()
            try:
                fw_proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                fw_proc.kill()

        # 8. temp-файл с URL'ами больше не нужен — удаляем
        try:
            os.remove(file_to_attach)
        except OSError as e:
            log.warning("не получилось удалить %s: %s", file_to_attach, e)

        log.info("сценарий завершён успешно")
        # Первые N успешных запусков подтверждаем push'ем — чтобы убедиться,
        # что setup на новой машине отработал. Дальше тишина (только падения).
        count = _read_success_count()
        if count < SUCCESS_NOTIFY_COUNT:
            count += 1
            _write_success_count(count)
            try:
                notify_ntfy(
                    _ntfy_header() +
                    f"UI install OK {count}/{SUCCESS_NOTIFY_COUNT}",
                    title="warmup OK",
                    priority="low",
                    tags="white_check_mark",
                )
            except Exception:
                pass
        return 0

    except Exception as exc:
        log.exception("сценарий упал: %s", exc)
        screenshot("ERROR", True)
        try:
            tail_lines: list[str] = []
            if LOG_FILE.exists():
                with LOG_FILE.open(encoding="utf-8", errors="ignore") as f:
                    tail_lines = f.readlines()[-15:]
            notify_ntfy(
                _ntfy_header() +
                f"error: {exc}\n\n"
                f"tail:\n" + "".join(tail_lines)
            )
        except Exception:
            pass
        return 1


if __name__ == "__main__":
    sys.exit(run())
